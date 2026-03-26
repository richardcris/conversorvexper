from __future__ import annotations

import hashlib
import json
import os
import queue
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from urllib.parse import urljoin, urlparse
from urllib.request import urlopen

import customtkinter as ctk
import firebirdsql
import xlrd
import xlwt
from openpyxl import Workbook as XlsxWorkbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageFilter


ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

APP_TITLE = "CONVERSOR - VEXPER"
APP_VERSION = "1.1.2"
WINDOW_SIZE = "1240x760"
PREVIEW_LIMIT = 100
EXPORT_BATCH_SIZE = 2000
PROGRESS_STEP = 2000
PREFERENCES_FILE = Path.home() / ".vexper_converter_prefs.json"
DEFAULT_LOGIN = {"username": "admin", "password": "1234"}
MODEL_TEMPLATES_DIRNAME = "PLANILHA MODELO"
XLS_MAX_ROWS = 65535
UPDATE_MANIFEST_NAME = "latest.json"
UPDATE_FEED_DIRNAME = "update_feed"
UPDATE_INSTALLER_NAME = "Instalador CONVERSOR - VEXPER.exe"
GITHUB_RELEASES_LATEST_SUFFIX = "/releases/latest/download/"

TEMPLATE_TABLE_ALIASES = {
    "TRANSPORTADORA": "transportadoras",
    "TRANSPORTADORAS": "transportadoras",
    "PRODUTOS_COMPOSICAO": "composicao",
    "COMPOSICAO": "composicao",
    "CLIENTES": "clientes",
    "PRODUTOS": "produtos",
    "GRADE": "grade",
    "FORNECEDOR": "fornecedores",
    "FORNECEDORES": "fornecedores",
}

HEADER_ALIASES: dict[str, dict[str, list[str]]] = {
    "PRODUTOS": {
        "CODIGO": ["ID_PRODUTO"],
        "BARRAS": ["BARRAS", "GTIN"],
        "BALANCA": ["CODIGO_BALANCA"],
        "DESCRICAO": ["PRODUTO", "DESCRICAO_COMPRA", "DESCRICAO_COMPLEMENTAR"],
        "GRUPO": ["GRUPO"],
        "MARCA": ["MARCA"],
        "REFERENCIA": ["REFERENCIA"],
        "SECAO": ["SECAO"],
        "LOCALIZACAO": ["LOCALIZACAO"],
        "APLICACAO": ["APLICACAO"],
        "CARACTERISTICA": ["CARACTERISTICA"],
        "TIPO_FATOR": ["TIPO_FATOR"],
        "FATOR": ["FATOR"],
        "ESTOQUE": ["ESTOQUE"],
        "ESTOQUE_MINIMO": ["MINIMO"],
        "UND": ["UNIDADE_COMECIAL"],
        "UND_ATACADO": ["UNIDADE_COMPRA"],
        "CUSTO": ["CUSTO", "VALOR_COMPRA"],
        "PRECO": ["VALOR_VENDA"],
        "ATACADO": ["VALOR_ATACADO"],
        "QTDE_PROMOCIONAL": ["QTDE_PROMOCIONAL"],
        "PROMOCAO": ["VALOR_PROMOCIONAL"],
        "MARGEM": ["MARGEM"],
        "NCM": ["NCM"],
        "CEST": ["CEST"],
        "ST": ["ST"],
        "ICMS": ["ICMS"],
        "RED_BASE_ICMS": ["RED_BASE_ICMS"],
        "MVA": ["MVA"],
        "STATUS": ["STATUS"],
        "NATUREZA_RECEITA_PISCOFINS": ["NATUREZA_RECEITA_PISCOFINS"],
        "CST_PIS_COFINS_ENTRADA": ["CST_PISCOFINS_ENTRADA"],
        "ALIQ_PIS_ENTRADA": ["ALIQ_PIS_ENTRADA"],
        "ALIQ_COFINS_ENTRADA": ["ALIQ_COFINS_ENTRADA"],
        "CST_PIS_COFINS_SAIDA": ["CST_PISCOFINS_SAIDA"],
        "ALIQ_PIS_SAIDA": ["ALIQ_PIS_SAIDA"],
        "ALIQ_COFINS_SAIDA": ["ALIQ_COFINS_SAIDA"],
    },
    "CLIENTES": {
        "CODIGO": ["ID_CLIENTE"],
        "CLIENTE": ["CLIENTE"],
        "FANTASIA": ["RAZ_SOCIAL", "CLIENTE"],
        "ENDERECO": ["LOGRADOURO"],
        "COMPLEMENTO": ["COMPLEMENTO"],
        "PONTO_REFERENCIA": ["PONTO_REFERENCIA"],
        "OBS": ["OBS"],
        "CIDADE": ["MUNICIPIO", "LOCALIDADE"],
        "BAIRRO": ["BAIRRO"],
        "CEP": ["CEP"],
        "UF": ["UF"],
        "NUMERO": ["NUMERO"],
        "TELEFONE": ["FONE"],
        "CELULAR": ["CELULAR"],
        "CELULAR2": ["CELULAR2"],
        "LIMITECREDITO": ["LMTE_CREDITO"],
        "CNPJ": ["CPF_CNPJ"],
        "CPF": ["CPF_CNPJ"],
        "IE": ["IE_RG"],
        "RG": ["IE_RG"],
        "PAI": ["PAI"],
        "MAE": ["MAE"],
        "EMAIL": ["EMAIL"],
        "CONTATO": ["CONTATO"],
    },
    "FORNECEDOR": {
        "CODIGO": ["ID_FORNECEDOR"],
        "RAZSOCIAL": ["RAZ_SOCIAL"],
        "FANTASIA": ["FANTASIA", "RAZ_SOCIAL"],
        "ENDERECO": ["ENDERECO"],
        "COMPLEMENTO": ["COMPLEMENTO"],
        "CIDADE": ["MUNICIPIO"],
        "BAIRRO": ["BAIRRO"],
        "CEP": ["CEP"],
        "UF": ["UF"],
        "NUMERO": ["NUMERO"],
        "TELEFONE": ["FONE"],
        "CELULAR": ["CELULAR"],
        "CNPJ": ["CNPJ"],
        "IE": ["IE"],
        "EMAIL": ["EMAIL"],
        "CONTATO": ["CONTATO"],
    },
    "FORNECEDORES": {
        "CODIGO": ["ID_FORNECEDOR"],
        "RAZSOCIAL": ["RAZ_SOCIAL"],
        "FANTASIA": ["FANTASIA", "RAZ_SOCIAL"],
        "ENDERECO": ["ENDERECO"],
        "COMPLEMENTO": ["COMPLEMENTO"],
        "CIDADE": ["MUNICIPIO"],
        "BAIRRO": ["BAIRRO"],
        "CEP": ["CEP"],
        "UF": ["UF"],
        "NUMERO": ["NUMERO"],
        "TELEFONE": ["FONE"],
        "CELULAR": ["CELULAR"],
        "CNPJ": ["CNPJ"],
        "IE": ["IE"],
        "EMAIL": ["EMAIL"],
        "CONTATO": ["CONTATO"],
    },
    "TRANSPORTADORA": {
        "CODIGO": ["ID", "ID_TRANSPORTADORA", "ID_TRANSPORT"],
        "RAZSOCIAL": ["RAZ_SOCIAL", "TRANSPORTADORA", "NOME"],
        "ENDERECO": ["LOGRADOURO", "ENDERECO"],
        "BAIRRO": ["BAIRRO"],
        "CIDADE": ["MUNICIPIO", "CIDADE"],
        "CEP": ["CEP"],
        "UF": ["UF"],
        "TELEFONE": ["FONE", "TELEFONE"],
        "CNPJ": ["CPF_CNPJ", "CNPJ"],
        "IE": ["IE_RG", "IE"],
    },
    "TRANSPORTADORAS": {
        "CODIGO": ["ID", "ID_TRANSPORTADORA", "ID_TRANSPORT"],
        "RAZSOCIAL": ["RAZ_SOCIAL", "TRANSPORTADORA", "NOME"],
        "ENDERECO": ["LOGRADOURO", "ENDERECO"],
        "BAIRRO": ["BAIRRO"],
        "CIDADE": ["MUNICIPIO", "CIDADE"],
        "CEP": ["CEP"],
        "UF": ["UF"],
        "TELEFONE": ["FONE", "TELEFONE"],
        "CNPJ": ["CPF_CNPJ", "CNPJ"],
        "IE": ["IE_RG", "IE", "INSC_EST"],
    },
    "GRADE": {
        "CODIGO": ["ID", "ID_GRADE", "ID_PRODUTOS_GRADE_ITENS"],
        "BARRAS": ["BARRAS", "GTIN"],
        "NOME_GRADE": ["GRADE", "NOME_GRADE", "DESCRICAO", "PRODUTO"],
        "ESTOQUE": ["ESTOQUE"],
    },
    "PRODUTOS_COMPOSICAO": {
        "ID_PRODUTO_COMPOSTO": ["ID_PRODUTO_COMPOSTO"],
        "ID_PRODUTO": ["ID_PRODUTO"],
        "DT_INSERT": ["DT_INSERT", "DT_CADASTRO"],
        "QUANTIDADE": ["QUANTIDADE", "QTDE"],
        "TIPO": ["TIPO", "TIPO_CADASTRO"],
    },
    "COMPOSICAO": {
        "ID_PRODUTO_COMPOSTO": ["ID_PRODUTO_COMPOSTO"],
        "ID_PRODUTO": ["ID_PRODUTO"],
        "DT_INSERT": ["DT_INSERT", "DT_CADASTRO"],
        "QUANTIDADE": ["QUANTIDADE", "QTDE"],
        "TIPO": ["TIPO", "TIPO_CADASTRO"],
    },
}

LOOKUP_ALIASES: dict[tuple[str, str], tuple[str, str, str]] = {
    ("PRODUTOS", "GRUPO"): ("PRODUTOS_GRUPO", "ID", "GRUPO"),
    ("PRODUTOS", "MARCA"): ("PRODUTOS_MARCA", "ID", "MARCA"),
}

THEMES = {
    "Oceano": {
        "app_bg": "#EAF2F8",
        "surface": "#F8FBFF",
        "surface_alt": "#EEF6FB",
        "header": "#07141F",
        "header_alt": "#0D2538",
        "text": "#183A5A",
        "muted": "#486581",
        "accent": "#2A9D8F",
        "accent_hover": "#22867A",
        "secondary": "#183A5A",
        "secondary_hover": "#102A43",
        "warm": "#F4A261",
        "warm_hover": "#E78D42",
        "badge": "#76E4F7",
        "list_bg": "#FFFFFF",
        "border": "#113554",
    },
    "Grafite": {
        "app_bg": "#E7EAEE",
        "surface": "#F8F9FB",
        "surface_alt": "#ECEFF3",
        "header": "#1D2128",
        "header_alt": "#2C3440",
        "text": "#273444",
        "muted": "#5B6675",
        "accent": "#00A896",
        "accent_hover": "#00897B",
        "secondary": "#374151",
        "secondary_hover": "#1F2937",
        "warm": "#F59E0B",
        "warm_hover": "#D97706",
        "badge": "#8EF6E4",
        "list_bg": "#FFFFFF",
        "border": "#3A4553",
    },
    "Aurora": {
        "app_bg": "#F3F7F1",
        "surface": "#FEFFFB",
        "surface_alt": "#F1F7EE",
        "header": "#132A13",
        "header_alt": "#244B2B",
        "text": "#1B4332",
        "muted": "#52796F",
        "accent": "#52B788",
        "accent_hover": "#40916C",
        "secondary": "#2D6A4F",
        "secondary_hover": "#1B4332",
        "warm": "#F4D35E",
        "warm_hover": "#EE964B",
        "badge": "#D8F3DC",
        "list_bg": "#FFFFFF",
        "border": "#2D6A4F",
    },
    "Solar": {
        "app_bg": "#FFF5EB",
        "surface": "#FFFDF8",
        "surface_alt": "#FFF1DE",
        "header": "#3A200D",
        "header_alt": "#5C3418",
        "text": "#7C2D12",
        "muted": "#9A5B2E",
        "accent": "#E76F51",
        "accent_hover": "#D95D39",
        "secondary": "#8D3C1F",
        "secondary_hover": "#6B2C15",
        "warm": "#F4A261",
        "warm_hover": "#E76F51",
        "badge": "#FFD6A5",
        "list_bg": "#FFFFFF",
        "border": "#8D3C1F",
    },
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="183A5A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)


def resource_path(name: str) -> Path:
    base_path = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base_path / name


def application_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def default_update_feed() -> str:
    env_feed = os.environ.get("VEXPER_UPDATE_FEED", "").strip()
    if env_feed:
        return env_feed

    github_repo = os.environ.get("VEXPER_GITHUB_REPO", "").strip().strip("/")
    if github_repo:
        return f"https://github.com/{github_repo}{GITHUB_RELEASES_LATEST_SUFFIX}"

    return str((application_base_dir() / UPDATE_FEED_DIRNAME).resolve())


def is_remote_source(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"}


def version_tuple(version: str) -> tuple[int, ...]:
    parts = []
    for chunk in re.findall(r"\d+", version):
        parts.append(int(chunk))
    return tuple(parts or [0])


def is_newer_version(candidate: str, current: str) -> bool:
    return version_tuple(candidate) > version_tuple(current)


def resolve_manifest_source(feed: str) -> str:
    if is_remote_source(feed):
        return urljoin(feed.rstrip("/") + "/", UPDATE_MANIFEST_NAME)
    path = Path(feed)
    if path.is_dir() or not path.suffix:
        return str(path / UPDATE_MANIFEST_NAME)
    return str(path)


def read_text_source(source: str) -> str:
    if is_remote_source(source):
        with urlopen(source, timeout=20) as response:
            return response.read().decode("utf-8")
    return Path(source).read_text(encoding="utf-8")


def resolve_installer_source(feed: str, location: str) -> str:
    if is_remote_source(location):
        return location
    if is_remote_source(feed):
        return urljoin(feed.rstrip("/") + "/", location)
    return str((Path(feed) / location).resolve())


def fetch_binary_source(source: str, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    if is_remote_source(source):
        with urlopen(source, timeout=60) as response:
            target.write_bytes(response.read())
        return
    shutil.copy2(source, target)


def load_update_manifest(feed: str) -> dict[str, object]:
    manifest_source = resolve_manifest_source(feed)
    content = read_text_source(manifest_source)
    data = json.loads(content)
    if not isinstance(data, dict):
        raise ValueError("Manifesto de atualizacao invalido.")
    return data


def load_preferences() -> dict[str, object]:
    defaults: dict[str, object] = {
        "theme": "Oceano",
        "sound_enabled": True,
        "auto_scan": True,
        "auto_update_enabled": True,
        "update_feed": default_update_feed(),
        "remember_login": False,
        "saved_username": "",
        "saved_password": "",
        "preview_limit": PREVIEW_LIMIT,
    }
    if not PREFERENCES_FILE.exists():
        return defaults

    try:
        loaded = json.loads(PREFERENCES_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return defaults

    if not isinstance(loaded, dict):
        return defaults
    defaults.update(loaded)
    return defaults


def save_preferences(preferences: dict[str, object]) -> None:
    PREFERENCES_FILE.write_text(json.dumps(preferences, indent=2, ensure_ascii=True), encoding="utf-8")


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def play_completion_sound() -> None:
    try:
        import winsound

        winsound.PlaySound("SystemExclamation", winsound.SND_ALIAS | winsound.SND_ASYNC)
    except Exception:
        pass


def load_logo_rgba() -> Image.Image | None:
    logo_path = resource_path("logo.png")
    if not logo_path.exists():
        return None

    logo = Image.open(logo_path).convert("RGBA")
    cleaned_pixels = []
    changed = False
    for red, green, blue, alpha in logo.getdata():
        if alpha > 0 and red <= 24 and green <= 24 and blue <= 24:
            cleaned_pixels.append((red, green, blue, 0))
            changed = True
        else:
            cleaned_pixels.append((red, green, blue, alpha))

    if changed:
        logo.putdata(cleaned_pixels)
    return logo


def build_glow_logo(size: tuple[int, int]) -> Image.Image | None:
    logo = load_logo_rgba()
    if logo is None:
        return None

    logo.thumbnail(size, Image.LANCZOS)

    canvas_size = (logo.width + 84, logo.height + 84)
    canvas = Image.new("RGBA", canvas_size, (0, 0, 0, 0))
    offset = ((canvas.width - logo.width) // 2, (canvas.height - logo.height) // 2)
    alpha = logo.getchannel("A")

    for color, blur_radius, opacity in (
        ((0, 244, 255), 16, 150),
        ((30, 144, 255), 30, 115),
        ((0, 255, 163), 48, 70),
    ):
        glow_alpha = Image.new("L", canvas_size, 0)
        glow_alpha.paste(alpha, offset)
        glow_alpha = glow_alpha.filter(ImageFilter.GaussianBlur(blur_radius))
        glow = Image.new("RGBA", canvas_size, color + (0,))
        glow.putalpha(glow_alpha.point(lambda value, scale=opacity: min(255, int(value * scale / 255))))
        canvas = Image.alpha_composite(canvas, glow)

    canvas.alpha_composite(logo, offset)
    return canvas


class LoginWindow(ctk.CTkToplevel):
    def __init__(self, master: "App") -> None:
        super().__init__(master)
        self.master_app = master
        self.geometry("540x660+440+90")
        self.resizable(False, False)
        self.title("Acesso interno")
        self.configure(fg_color=master.theme["app_bg"])
        self.transient(master)
        self.attributes("-topmost", True)
        self.protocol("WM_DELETE_WINDOW", self._cancel)

        self.username_var = tk.StringVar(value=str(master.preferences.get("saved_username", "")))
        self.password_var = tk.StringVar(value=str(master.preferences.get("saved_password", "")))
        self.remember_var = tk.BooleanVar(value=bool(master.preferences.get("remember_login", False)))
        self.show_password_var = tk.BooleanVar(value=False)
        self._pulse = 0

        card = ctk.CTkFrame(self, corner_radius=26, fg_color=master.theme["surface"])
        card.pack(fill="both", expand=True, padx=22, pady=22)

        self.logo_label = ctk.CTkLabel(card, text="", image=master.splash_logo_image)
        self.logo_label.pack(pady=(18, 8))

        ctk.CTkLabel(
            card,
            text="Login interno",
            font=ctk.CTkFont(family="Segoe UI", size=28, weight="bold"),
            text_color=master.theme["text"],
        ).pack()

        ctk.CTkLabel(
            card,
            text="Entre para liberar o navegador de tabelas e a conversao para Excel.",
            font=ctk.CTkFont(family="Segoe UI", size=14),
            text_color=master.theme["muted"],
        ).pack(pady=(8, 14))

        self.pulse_bar = ctk.CTkProgressBar(card, width=320, progress_color=master.theme["accent"])
        self.pulse_bar.pack(pady=(0, 18))
        self.pulse_bar.set(0.25)

        self.username_entry = self._entry_block(card, "Usuario", self.username_var)
        self.password_entry = self._entry_block(card, "Senha", self.password_var, show="*")

        toggles = ctk.CTkFrame(card, fg_color="transparent")
        toggles.pack(fill="x", padx=34, pady=(2, 10))
        toggles.grid_columnconfigure((0, 1), weight=1)

        self.remember_check = ctk.CTkCheckBox(
            toggles,
            text="Salvar acesso",
            variable=self.remember_var,
            text_color=master.theme["text"],
            fg_color=master.theme["accent"],
            hover_color=master.theme["accent_hover"],
        )
        self.remember_check.grid(row=0, column=0, sticky="w")

        self.show_password_check = ctk.CTkCheckBox(
            toggles,
            text="Ver senha",
            variable=self.show_password_var,
            command=self._toggle_password,
            text_color=master.theme["text"],
            fg_color=master.theme["accent"],
            hover_color=master.theme["accent_hover"],
        )
        self.show_password_check.grid(row=0, column=1, sticky="e")

        self.feedback_label = ctk.CTkLabel(card, text="Use admin / 1234 no primeiro acesso.", text_color=master.theme["muted"])
        self.feedback_label.pack(pady=(0, 10))

        self.login_button = ctk.CTkButton(
            card,
            text="Entrar no sistema",
            height=46,
            corner_radius=16,
            fg_color=master.theme["accent"],
            hover_color=master.theme["accent_hover"],
            command=self._submit,
        )
        self.login_button.pack(fill="x", padx=34, pady=(0, 8))

        ctk.CTkButton(
            card,
            text="Fechar",
            height=40,
            corner_radius=16,
            fg_color=master.theme["surface_alt"],
            text_color=master.theme["text"],
            hover_color=master.theme["app_bg"],
            command=self._cancel,
        ).pack(fill="x", padx=34, pady=(0, 18))

        self.after(80, self._animate)
        self.bind("<Return>", lambda _event: self._submit())
        self.after(150, self._focus_username)

    def _entry_block(self, parent, label: str, variable: tk.StringVar, show: str | None = None):
        ctk.CTkLabel(parent, text=label, text_color=self.master_app.theme["muted"]).pack(anchor="w", padx=34, pady=(0, 6))
        entry = ctk.CTkEntry(parent, textvariable=variable, height=42, show=show or "")
        entry.pack(fill="x", padx=34, pady=(0, 14))
        return entry

    def _focus_username(self) -> None:
        if not self.winfo_exists():
            return
        self.lift()
        self.focus_force()
        self.username_entry.focus_force()

    def _animate(self) -> None:
        self._pulse = (self._pulse + 7) % 100
        value = 0.2 + (abs(50 - self._pulse) / 100)
        self.pulse_bar.set(value)
        if self.winfo_exists():
            self.after(80, self._animate)

    def _toggle_password(self) -> None:
        self.password_entry.configure(show="" if self.show_password_var.get() else "*")

    def _submit(self) -> None:
        username = self.username_var.get().strip()
        password = self.password_var.get()
        if username == DEFAULT_LOGIN["username"] and hash_password(password) == hash_password(DEFAULT_LOGIN["password"]):
            self.master_app.preferences["remember_login"] = self.remember_var.get()
            self.master_app.preferences["saved_username"] = username if self.remember_var.get() else ""
            self.master_app.preferences["saved_password"] = password if self.remember_var.get() else ""
            save_preferences(self.master_app.preferences)
            self.master_app.on_login_success(username)
            self.destroy()
            return

        self.feedback_label.configure(text="Usuario ou senha invalidos.", text_color="#D62828")

    def _cancel(self) -> None:
        self.master_app.destroy()


@dataclass(slots=True)
class TableProfile:
    name: str
    row_count: int
    column_count: int
    columns: list[str]


@dataclass(slots=True)
class ExportResult:
    output_path: Path
    profiles: list[TableProfile]


@dataclass(slots=True)
class UpdatePackage:
    version: str
    installer_source: str
    installer_name: str
    notes: str


def model_templates_dir() -> Path:
    return resource_path(MODEL_TEMPLATES_DIRNAME)


def canonical_table_name(table_name: str) -> str:
    normalized = table_name.strip().upper()
    return TEMPLATE_TABLE_ALIASES.get(normalized, normalized).upper()


def template_path_for_table(table_name: str) -> Path | None:
    normalized = TEMPLATE_TABLE_ALIASES.get(table_name.strip().upper(), table_name.strip().lower())
    candidate = model_templates_dir() / f"{normalized}.xls"
    if candidate.exists():
        return candidate
    return None


def load_template_headers(table_name: str) -> list[str] | None:
    template_path = template_path_for_table(table_name)
    if template_path is None:
        return None

    workbook = xlrd.open_workbook(template_path)
    sheet = workbook.sheet_by_index(0)
    return [str(value).strip() for value in sheet.row_values(0)]


def resolve_source_column(table_name: str, header_name: str, source_index: dict[str, int]) -> int | None:
    canonical = canonical_table_name(table_name)
    aliases = HEADER_ALIASES.get(canonical, {})
    candidates = aliases.get(header_name.upper(), [header_name.upper()])
    for candidate in candidates:
        position = source_index.get(candidate.upper())
        if position is not None:
            return position
    return None


class DatabaseExcelConverter:
    def __init__(
        self,
        database_path: Path,
        host: str = "127.0.0.1",
        port: int = 3050,
        user: str = "SYSDBA",
        password: str = "masterkey",
        charset: str = "WIN1252",
    ) -> None:
        self.database_path = database_path
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.charset = charset
        self.engine = self._detect_engine()
        self.lookup_cache: dict[tuple[str, str, str], dict[object, object]] = {}

    def _detect_engine(self) -> str:
        suffix = self.database_path.suffix.lower()
        if suffix in {".db", ".sqlite", ".sqlite3"}:
            return "sqlite"
        if suffix == ".fdb":
            return "firebird"
        raise ValueError("Formato de banco não suportado. Use SQLite (.db, .sqlite, .sqlite3) ou Firebird (.fdb).")

    def _connect(self):
        if self.engine == "sqlite":
            connection = sqlite3.connect(self.database_path)
            connection.execute("PRAGMA query_only = ON")
            return connection

        return firebirdsql.connect(
            host=self.host,
            port=self.port,
            database=str(self.database_path),
            user=self.user,
            password=self.password,
            charset=self.charset,
        )

    def list_tables(self) -> list[str]:
        with self._connect() as connection:
            cursor = connection.cursor()
            if self.engine == "sqlite":
                cursor.execute(
                    """
                    SELECT name
                    FROM sqlite_master
                    WHERE type = 'table'
                      AND name NOT LIKE 'sqlite_%'
                    ORDER BY name
                    """
                )
                rows = cursor.fetchall()
                cursor.close()
                return [str(row[0]) for row in rows]

            cursor.execute(
                """
                SELECT TRIM(rdb$relation_name)
                FROM rdb$relations
                WHERE rdb$view_blr IS NULL
                  AND COALESCE(rdb$system_flag, 0) = 0
                ORDER BY rdb$relation_name
                """
            )
            rows = cursor.fetchall()
            cursor.close()
            return [str(row[0]).strip() for row in rows]

    def read_preview(self, table_name: str, limit: int = PREVIEW_LIMIT) -> tuple[list[str], list[tuple]]:
        safe_name = escape_identifier(table_name)
        with self._connect() as connection:
            cursor = connection.cursor()
            if self.engine == "sqlite":
                cursor.execute(f'SELECT * FROM "{safe_name}" LIMIT ?', (limit,))
            else:
                cursor.execute(f'SELECT FIRST {int(limit)} * FROM "{safe_name}"')
            columns = [str(column[0]).strip() for column in cursor.description or []]
            rows = cursor.fetchall()
            cursor.close()
        return columns, rows

    def get_profile(self, table_name: str) -> TableProfile:
        columns, _ = self.read_preview(table_name, limit=1)
        safe_name = escape_identifier(table_name)
        with self._connect() as connection:
            cursor = connection.cursor()
            cursor.execute(f'SELECT COUNT(*) FROM "{safe_name}"')
            row_count = int(cursor.fetchone()[0])
            cursor.close()
        return TableProfile(
            name=table_name,
            row_count=row_count,
            column_count=len(columns),
            columns=columns,
        )

    def export_tables(self, table_names: list[str], output_path: Path, progress_callback) -> list[TableProfile]:
        if output_path.suffix.lower() == ".xls":
            return self._export_tables_xls(table_names, output_path, progress_callback)

        workbook = XlsxWorkbook(write_only=True)
        profiles: list[TableProfile] = []
        used_names: set[str] = set()

        with self._connect() as connection:
            total_tables = max(len(table_names), 1)
            for index, table_name in enumerate(table_names, start=1):
                profile = self._write_table_sheet(
                    workbook=workbook,
                    connection=connection,
                    table_name=table_name,
                    used_names=used_names,
                    progress_callback=progress_callback,
                    table_position=index,
                    total_tables=total_tables,
                )
                profiles.append(profile)

        self._write_summary_sheet(workbook, profiles, used_names)
        workbook.save(output_path)
        return profiles

    def _load_lookup_map(self, lookup_table: str, key_column: str, value_column: str) -> dict[object, object]:
        cache_key = (lookup_table, key_column, value_column)
        if cache_key in self.lookup_cache:
            return self.lookup_cache[cache_key]

        safe_table = escape_identifier(lookup_table)
        safe_key = escape_identifier(key_column)
        safe_value = escape_identifier(value_column)
        with self._connect() as connection:
            cursor = connection.cursor()
            cursor.execute(f'SELECT "{safe_key}", "{safe_value}" FROM "{safe_table}"')
            rows = cursor.fetchall()
            cursor.close()

        lookup = {normalize_cell_value(row[0]): normalize_cell_value(row[1]) for row in rows}
        self.lookup_cache[cache_key] = lookup
        return lookup

    def _resolve_export_value(self, table_name: str, header_name: str, row: tuple, source_pos: int | None):
        raw_value = ""
        if source_pos is not None and source_pos < len(row):
            raw_value = normalize_cell_value(row[source_pos])

        lookup_key = (canonical_table_name(table_name), header_name.upper())
        lookup_spec = LOOKUP_ALIASES.get(lookup_key)
        if lookup_spec is not None and raw_value not in (None, ""):
            lookup = self._load_lookup_map(*lookup_spec)
            return lookup.get(raw_value, raw_value)

        return raw_value

    def _export_tables_xls(self, table_names: list[str], output_path: Path, progress_callback) -> list[TableProfile]:
        workbook = xlwt.Workbook(encoding="utf-8")
        profiles: list[TableProfile] = []
        used_names: set[str] = set()

        with self._connect() as connection:
            total_tables = max(len(table_names), 1)
            for index, table_name in enumerate(table_names, start=1):
                profile = self._write_table_sheet_xls(
                    workbook=workbook,
                    connection=connection,
                    table_name=table_name,
                    used_names=used_names,
                    progress_callback=progress_callback,
                    table_position=index,
                    total_tables=total_tables,
                )
                profiles.append(profile)

        workbook.save(str(output_path))
        return profiles

    def _write_table_sheet(
        self,
        workbook: XlsxWorkbook,
        connection,
        table_name: str,
        used_names: set[str],
        progress_callback,
        table_position: int,
        total_tables: int,
    ) -> TableProfile:
        safe_name = escape_identifier(table_name)
        cursor = connection.cursor()
        if hasattr(cursor, "arraysize"):
            cursor.arraysize = EXPORT_BATCH_SIZE

        cursor.execute(f'SELECT * FROM "{safe_name}"')
        columns = [str(column[0]).strip() for column in cursor.description or []]

        worksheet = workbook.create_sheet(title=sanitize_sheet_name(table_name, used_names))
        worksheet.freeze_panes = "A2"
        write_header_row(worksheet, columns)
        apply_fast_layout(worksheet, columns)

        row_count = 0
        while True:
            batch = cursor.fetchmany(EXPORT_BATCH_SIZE)
            if not batch:
                break

            for row in batch:
                worksheet.append([normalize_cell_value(value) for value in row])
                row_count += 1

            if row_count and row_count % PROGRESS_STEP == 0:
                table_progress = ((table_position - 1) + 0.75) / total_tables
                progress_callback(
                    f"Lendo {table_name}: {row_count} linha(s) processadas",
                    min(0.92, table_progress),
                )

        cursor.close()

        if columns:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, row_count + 1)}"

        progress_callback(
            f"Tabela {table_name} exportada com {row_count} linha(s)",
            min(0.96, table_position / total_tables),
        )
        return TableProfile(
            name=table_name,
            row_count=row_count,
            column_count=len(columns),
            columns=columns,
        )

    def _write_table_sheet_xls(
        self,
        workbook: xlwt.Workbook,
        connection,
        table_name: str,
        used_names: set[str],
        progress_callback,
        table_position: int,
        total_tables: int,
    ) -> TableProfile:
        safe_name = escape_identifier(table_name)
        cursor = connection.cursor()
        if hasattr(cursor, "arraysize"):
            cursor.arraysize = EXPORT_BATCH_SIZE

        cursor.execute(f'SELECT * FROM "{safe_name}"')
        source_columns = [str(column[0]).strip() for column in cursor.description or []]
        template_headers = load_template_headers(table_name)
        export_columns = template_headers or source_columns
        source_index = {column.upper(): idx for idx, column in enumerate(source_columns)}

        sheet_name = sanitize_sheet_name("Planilha1" if template_headers else table_name, used_names)
        worksheet = workbook.add_sheet(sheet_name)
        header_style = build_xls_header_style()
        cell_style = build_xls_cell_style()
        date_style = build_xls_date_style()

        for col_index, column_name in enumerate(export_columns):
            worksheet.write(0, col_index, column_name, header_style)
            worksheet.col(col_index).width = min(max((len(str(column_name)) + 2) * 256, 14 * 256), 28 * 256)

        row_count = 0
        while True:
            batch = cursor.fetchmany(EXPORT_BATCH_SIZE)
            if not batch:
                break

            for row in batch:
                target_row = row_count + 1
                if target_row > XLS_MAX_ROWS:
                    cursor.close()
                    raise ValueError(
                        f"A tabela {table_name} possui mais de {XLS_MAX_ROWS} linhas e nao cabe em .xls. Use .xlsx para essa exportacao."
                    )

                for col_index, column_name in enumerate(export_columns):
                    source_pos = resolve_source_column(table_name, column_name, source_index)
                    raw_value = self._resolve_export_value(table_name, column_name, row, source_pos)
                    write_xls_value(worksheet, target_row, col_index, raw_value, cell_style, date_style)

                row_count += 1

            if row_count and row_count % PROGRESS_STEP == 0:
                table_progress = ((table_position - 1) + 0.75) / total_tables
                progress_callback(
                    f"Lendo {table_name}: {row_count} linha(s) processadas",
                    min(0.92, table_progress),
                )

        cursor.close()
        progress_callback(
            f"Tabela {table_name} exportada no modelo .xls com {row_count} linha(s)",
            min(0.96, table_position / total_tables),
        )
        return TableProfile(
            name=table_name,
            row_count=row_count,
            column_count=len(export_columns),
            columns=export_columns,
        )

    def _write_summary_sheet(self, workbook: XlsxWorkbook, profiles: list[TableProfile], used_names: set[str]) -> None:
        worksheet = workbook.create_sheet(title=sanitize_sheet_name("Resumo", used_names))
        headers = ["Tabela", "Linhas", "Colunas", "Campos"]
        worksheet.freeze_panes = "A2"
        write_header_row(worksheet, headers)
        apply_fast_layout(worksheet, headers)

        for profile in profiles:
            worksheet.append([profile.name, profile.row_count, profile.column_count, ", ".join(profile.columns)])

        if headers:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(profiles) + 1)}"


def escape_identifier(name: str) -> str:
    return name.replace('"', '""')


def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "Planilha"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix_index = 1

    while candidate in used_names:
        suffix = f"_{suffix_index}"
        candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
        suffix_index += 1

    used_names.add(candidate)
    return candidate


def build_output_path(database_path: Path) -> Path:
    candidate = database_path.with_name(f"{database_path.stem}_convertido.xlsx")
    if not candidate.exists():
        return candidate

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return database_path.with_name(f"{database_path.stem}_convertido_{timestamp}.xlsx")


def normalize_cell_value(value):
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return value


def build_xls_header_style() -> xlwt.XFStyle:
    style = xlwt.XFStyle()
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 55
    style.pattern = pattern

    font = xlwt.Font()
    font.bold = True
    font.colour_index = 1
    style.font = font

    alignment = xlwt.Alignment()
    alignment.horz = xlwt.Alignment.HORZ_CENTER
    alignment.vert = xlwt.Alignment.VERT_CENTER
    style.alignment = alignment
    return style


def build_xls_cell_style() -> xlwt.XFStyle:
    style = xlwt.XFStyle()
    alignment = xlwt.Alignment()
    alignment.horz = xlwt.Alignment.HORZ_LEFT
    alignment.vert = xlwt.Alignment.VERT_CENTER
    style.alignment = alignment
    return style


def build_xls_date_style() -> xlwt.XFStyle:
    style = build_xls_cell_style()
    style.num_format_str = "dd/mm/yyyy"
    return style


def write_xls_value(worksheet, row_index: int, col_index: int, value, default_style: xlwt.XFStyle, date_style: xlwt.XFStyle) -> None:
    if value is None:
        worksheet.write(row_index, col_index, "", default_style)
        return
    if isinstance(value, datetime):
        worksheet.write(row_index, col_index, value, date_style)
        return
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day") and not isinstance(value, str):
        worksheet.write(row_index, col_index, value, date_style)
        return
    worksheet.write(row_index, col_index, value, default_style)


def write_header_row(worksheet, columns: list[str]) -> None:
    header_row = []
    for column_name in columns:
        cell = WriteOnlyCell(worksheet, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        header_row.append(cell)
    worksheet.append(header_row)


def apply_fast_layout(worksheet, columns: list[str]) -> None:
    for column_index, column_name in enumerate(columns, start=1):
        adjusted_width = min(max(len(str(column_name)) + 4, 14), 26)
        worksheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


class SplashScreen(ctk.CTkToplevel):
    def __init__(self, master, logo_image=None) -> None:
        super().__init__(master)
        self.overrideredirect(True)
        self.geometry("620x360+390+170")
        self.attributes("-topmost", True)
        self.configure(fg_color="#07141F")
        self.attributes("-alpha", 0.0)

        self.logo_label = ctk.CTkLabel(self, text="", image=logo_image)
        self.logo_label.pack(pady=(34, 10))

        self.title_label = ctk.CTkLabel(
            self,
            text=APP_TITLE,
            font=ctk.CTkFont(family="Segoe UI", size=34, weight="bold"),
            text_color="#F4F7FB",
        )
        self.title_label.pack(pady=(0, 12))

        self.subtitle_label = ctk.CTkLabel(
            self,
            text="Leitura rapida de bancos e exportacao organizada para Excel",
            font=ctk.CTkFont(family="Segoe UI", size=16),
            text_color="#B8C7D9",
        )
        self.subtitle_label.pack()

        self.progress = ctk.CTkProgressBar(self, width=360, progress_color="#2A9D8F")
        self.progress.pack(pady=(42, 10))
        self.progress.set(0)

        self.status = ctk.CTkLabel(
            self,
            text="Inicializando interface...",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color="#DCE6F2",
        )
        self.status.pack()

        self._alpha = 0.0
        self._progress_value = 0.0
        self.after(20, self.animate)

    def animate(self) -> None:
        self._alpha = min(1.0, self._alpha + 0.06)
        self._progress_value = min(1.0, self._progress_value + 0.04)
        self.attributes("-alpha", self._alpha)
        self.progress.set(self._progress_value)

        if self._progress_value < 1.0:
            self.after(35, self.animate)
            return

        self.status.configure(text="Interface pronta")
        self.after(350, self.destroy)


class App(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        self.preferences = load_preferences()
        self.theme_name = str(self.preferences.get("theme", "Oceano"))
        self.theme = THEMES.get(self.theme_name, THEMES["Oceano"])
        self.title(APP_TITLE)
        self.geometry(WINDOW_SIZE)
        self.minsize(1100, 680)
        self.configure(fg_color=self.theme["app_bg"])

        self.database_path: Path | None = None
        self.converter: DatabaseExcelConverter | None = None
        self.table_profiles: dict[str, TableProfile] = {}
        self.all_tables: list[str] = []
        self.filtered_tables: list[str] = []
        self.progress_queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.scan_thread: threading.Thread | None = None
        self.profile_thread: threading.Thread | None = None
        self.preview_thread: threading.Thread | None = None
        self.export_thread: threading.Thread | None = None
        self.update_thread: threading.Thread | None = None
        self.connection_panel_visible = False
        self.window_icon = None
        self.header_logo_image = None
        self.splash_logo_image = None
        self.login_window: LoginWindow | None = None
        self.settings_window: ctk.CTkToplevel | None = None
        self.current_user = ""
        self.pending_update: UpdatePackage | None = None

        self._load_brand_assets()
        self._apply_window_icon()

        self.withdraw()
        self.splash = SplashScreen(self, self.splash_logo_image)
        self.after(900, self._show_login_window)

        self._build_layout()
        self._apply_theme()
        self.after(150, self._poll_queue)

    def _load_brand_assets(self) -> None:
        header_logo = build_glow_logo((220, 82))
        if header_logo is not None:
            self.header_logo_image = ctk.CTkImage(light_image=header_logo, dark_image=header_logo, size=header_logo.size)

        splash_logo = build_glow_logo((280, 104))
        if splash_logo is not None:
            self.splash_logo_image = ctk.CTkImage(light_image=splash_logo, dark_image=splash_logo, size=splash_logo.size)

    def _apply_window_icon(self) -> None:
        logo_path = resource_path("logo.png")
        if not logo_path.exists():
            return
        try:
            self.window_icon = tk.PhotoImage(file=str(logo_path))
            self.iconphoto(True, self.window_icon)
        except tk.TclError:
            self.window_icon = None

    def _show_login_window(self) -> None:
        self.login_window = LoginWindow(self)
        self.login_window.grab_set()
        self._start_update_check()

    def on_login_success(self, username: str) -> None:
        self.current_user = username
        self.deiconify()
        self.user_label.configure(text=f"Usuario: {username}")
        self.status_label.configure(text=f"Acesso liberado para {username}. Selecione um banco para iniciar.")
        if self.pending_update is not None:
            self.after(250, lambda: self._prompt_update_package(self.pending_update))

    def _start_update_check(self) -> None:
        if not bool(self.preferences.get("auto_update_enabled", True)):
            return
        if self.update_thread and self.update_thread.is_alive():
            return
        self.update_thread = threading.Thread(target=self._run_update_check, daemon=True)
        self.update_thread.start()

    def _run_update_check(self) -> None:
        feed = str(self.preferences.get("update_feed", "")).strip()
        if not feed:
            return

        try:
            manifest = load_update_manifest(feed)
            latest_version = str(manifest.get("version", "")).strip()
            if not latest_version or not is_newer_version(latest_version, APP_VERSION):
                return

            installer_name = str(manifest.get("installer_name", UPDATE_INSTALLER_NAME)).strip() or UPDATE_INSTALLER_NAME
            installer_location = str(manifest.get("installer_location", installer_name)).strip() or installer_name
            notes = str(manifest.get("notes", "Atualizacao disponivel.")).strip()
            installer_source = resolve_installer_source(feed, installer_location)

            temp_dir = Path(tempfile.gettempdir()) / "vexper_converter_update"
            target = temp_dir / installer_name
            self.progress_queue.put(("update_progress", f"Baixando atualizacao {latest_version}..."))
            fetch_binary_source(installer_source, target)
            package = UpdatePackage(version=latest_version, installer_source=str(target), installer_name=installer_name, notes=notes)
            self.progress_queue.put(("update_ready", package))
        except Exception as error:
            self.progress_queue.put(("update_error", str(error)))

    def _install_update_package(self, package: UpdatePackage) -> None:
        installer_path = Path(package.installer_source)
        if not installer_path.exists():
            messagebox.showerror("Atualizacao", "O instalador baixado nao foi encontrado.")
            return

        updater_script = installer_path.with_suffix(".cmd")
        updater_script.write_text(
            "@echo off\n"
            "ping 127.0.0.1 -n 3 > nul\n"
            f'start "" "{installer_path}" /SP- /VERYSILENT /SUPPRESSMSGBOXES /NORESTART /CLOSEAPPLICATIONS /FORCECLOSEAPPLICATIONS\n',
            encoding="utf-8",
        )
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        subprocess.Popen(["cmd", "/c", str(updater_script)], creationflags=creationflags)
        self.after(300, self.destroy)

    def _prompt_update_package(self, package: UpdatePackage | None) -> None:
        if package is None:
            return
        if not self.current_user:
            return

        answer = messagebox.askyesno(
            "Atualizacao disponivel",
            f"Foi encontrada a versao {package.version}.\n\n{package.notes}\n\nDeseja atualizar agora?",
            parent=self,
        )
        if answer:
            self.status_label.configure(text=f"Atualizacao {package.version} encontrada. Instalando nova versao...")
            self._install_update_package(package)
            return

        self.status_label.configure(text=f"Atualizacao {package.version} identificada. Voce pode atualizar depois nas proximas aberturas do sistema.")

    def _build_layout(self) -> None:
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.header = ctk.CTkFrame(self, corner_radius=0, fg_color=self.theme["header"], height=104)
        self.header.grid(row=0, column=0, columnspan=2, sticky="nsew")
        self.header.grid_columnconfigure(0, weight=1)
        self.header.grid_columnconfigure(1, weight=0)

        self.header_text_frame = ctk.CTkFrame(self.header, fg_color="transparent")
        self.header_text_frame.grid(row=0, column=0, padx=(28, 0), pady=20, sticky="w")

        self.badge_label = ctk.CTkLabel(
            self.header_text_frame,
            text="VEXPER SISTEMAS",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=self.theme["badge"],
            fg_color=self.theme["header_alt"],
            corner_radius=999,
            padx=14,
            pady=4,
        )
        self.badge_label.pack(anchor="w", pady=(0, 6))

        self.version_label = ctk.CTkLabel(
            self.header_text_frame,
            text=f"Versao {APP_VERSION}",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color="#DCE6F2",
            fg_color=self.theme["header_alt"],
            corner_radius=999,
            padx=12,
            pady=3,
        )
        self.version_label.pack(anchor="w", pady=(0, 8))

        title = ctk.CTkLabel(
            self.header_text_frame,
            text=APP_TITLE,
            font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"),
            text_color="#F8FBFF",
        )
        title.pack(anchor="w")

        subtitle = ctk.CTkLabel(
            self.header_text_frame,
            text="Abra o banco, visualize as tabelas e exporte com rapidez para Excel.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color="#C6D3E1",
        )
        subtitle.pack(anchor="w", pady=(4, 0))

        self.header_actions = ctk.CTkFrame(self.header, fg_color="transparent")
        self.header_actions.grid(row=0, column=1, padx=(12, 26), pady=16, sticky="e")

        self.user_label = ctk.CTkLabel(
            self.header_actions,
            text="Aguardando login",
            text_color="#DCE6F2",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
        )
        self.user_label.pack(anchor="e", pady=(0, 8))

        self.settings_button = ctk.CTkButton(
            self.header_actions,
            text="Configuracao",
            width=140,
            height=36,
            corner_radius=12,
            fg_color=self.theme["surface_alt"],
            text_color=self.theme["text"],
            hover_color=self.theme["app_bg"],
            command=self.open_settings,
        )
        self.settings_button.pack(anchor="e")

        self.brand_glow_label = ctk.CTkLabel(self.header, text="", image=self.header_logo_image, fg_color="transparent")
        self.brand_glow_label.place(relx=0.64, rely=0.5, anchor="center")

        self.sidebar = ctk.CTkFrame(self, width=332, corner_radius=20, fg_color=self.theme["surface"])
        self.sidebar.grid(row=1, column=0, padx=(18, 10), pady=14, sticky="nsew")
        self.sidebar.grid_propagate(False)

        self.content = ctk.CTkFrame(self, corner_radius=20, fg_color=self.theme["surface"])
        self.content.grid(row=1, column=1, padx=(10, 18), pady=14, sticky="nsew")
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(3, weight=1)

        self._build_sidebar()
        self._build_content()

    def _build_sidebar(self) -> None:
        ctk.CTkLabel(
            self.sidebar,
            text="Arquivo",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color="#183A5A",
        ).pack(anchor="w", padx=22, pady=(22, 8))

        self.pick_button = ctk.CTkButton(
            self.sidebar,
            text="Selecionar banco de dados",
            height=44,
            corner_radius=14,
            fg_color="#2A9D8F",
            hover_color="#22867A",
            command=self.pick_database,
        )
        self.pick_button.pack(fill="x", padx=22, pady=(0, 12))

        self.file_label = ctk.CTkLabel(
            self.sidebar,
            text="Nenhum arquivo selecionado",
            justify="left",
            wraplength=280,
            text_color="#486581",
        )
        self.file_label.pack(anchor="w", padx=22)

        self.connection_toggle_button = ctk.CTkButton(
            self.sidebar,
            text="Mostrar conexão Firebird",
            height=34,
            corner_radius=12,
            fg_color="#D9EAF7",
            text_color="#183A5A",
            hover_color="#C4DDF0",
            command=self._toggle_connection_panel,
        )
        self.connection_toggle_button.pack(fill="x", padx=22, pady=(16, 10))

        self.connection_frame = ctk.CTkFrame(self.sidebar, fg_color="#F4F8FB", corner_radius=16)

        ctk.CTkLabel(
            self.connection_frame,
            text="Conexao Firebird (.FDB)",
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
            text_color="#183A5A",
        ).pack(anchor="w", padx=16, pady=(14, 8))

        self.host_entry = self._sidebar_entry("Host", "127.0.0.1", parent=self.connection_frame)
        self.port_entry = self._sidebar_entry("Porta", "3050", parent=self.connection_frame)
        self.user_entry = self._sidebar_entry("Usuario", "SYSDBA", parent=self.connection_frame)
        self.password_entry = self._sidebar_entry("Senha", "masterkey", parent=self.connection_frame, show="*")
        self.charset_entry = self._sidebar_entry("Charset", "WIN1252", parent=self.connection_frame)

        self.scan_button = ctk.CTkButton(
            self.sidebar,
            text="Ler estrutura do banco",
            height=42,
            corner_radius=14,
            fg_color="#183A5A",
            hover_color="#102A43",
            command=self.scan_database,
            state="disabled",
        )
        self.scan_button.pack(fill="x", padx=22, pady=(18, 20))

        ctk.CTkLabel(
            self.sidebar,
            text="Tabelas",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color="#183A5A",
        ).pack(anchor="w", padx=22)

        self.table_search_var = tk.StringVar()
        self.table_search_var.trace_add("write", self._on_table_search_changed)

        self.table_search_entry = ctk.CTkEntry(
            self.sidebar,
            height=36,
            textvariable=self.table_search_var,
            placeholder_text="Buscar tabela, ex: PRODUTOS",
        )
        self.table_search_entry.pack(fill="x", padx=22, pady=(10, 10))

        self.list_frame = tk.Frame(self.sidebar, bg=self.theme["surface"])
        self.list_frame.pack(fill="both", expand=True, padx=22, pady=(0, 12))

        self.table_listbox = tk.Listbox(
            self.list_frame,
            selectmode=tk.BROWSE,
            bg=self.theme["list_bg"],
            fg=self.theme["text"],
            highlightthickness=0,
            borderwidth=0,
            font=("Segoe UI", 11),
            activestyle="none",
            exportselection=False,
        )
        self.table_listbox.pack(side="left", fill="both", expand=True)
        self.table_listbox.bind("<<ListboxSelect>>", self.on_table_selected)

        self.table_scrollbar = tk.Scrollbar(
            self.list_frame,
            orient="vertical",
            command=self.table_listbox.yview,
            width=18,
            activebackground=self.theme["accent"],
            bg=self.theme["surface_alt"],
            troughcolor=self.theme["app_bg"],
            relief="flat",
            bd=0,
        )
        self.table_scrollbar.pack(side="right", fill="y")
        self.table_listbox.configure(yscrollcommand=self.table_scrollbar.set)
        self.table_listbox.bind("<MouseWheel>", self._on_listbox_mousewheel)

        self.select_all_button = ctk.CTkButton(
            self.sidebar,
            text="Selecionar todas",
            height=38,
            corner_radius=12,
            fg_color="#D9EAF7",
            text_color="#183A5A",
            hover_color="#C4DDF0",
            command=self.select_all_tables,
            state="disabled",
        )
        self.select_all_button.pack(fill="x", padx=22, pady=(0, 8))

        self.export_all_button = ctk.CTkButton(
            self.sidebar,
            text="Exportar banco inteiro",
            height=42,
            corner_radius=14,
            fg_color="#F4A261",
            hover_color="#E78D42",
            text_color="#1F2933",
            command=self.export_all_tables,
            state="disabled",
        )
        self.export_all_button.pack(fill="x", padx=22, pady=(0, 8))

    def _build_content(self) -> None:
        self.metrics_frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.metrics_frame.grid(row=0, column=0, padx=24, pady=(24, 12), sticky="ew")
        self.metrics_frame.grid_columnconfigure((0, 1, 2), weight=1)

        self.metric_tables = self._metric_card(self.metrics_frame, "Tabelas", "0", 0)
        self.metric_rows = self._metric_card(self.metrics_frame, "Linhas", "0", 1)
        self.metric_columns = self._metric_card(self.metrics_frame, "Colunas", "0", 2)

        self.progress_frame = ctk.CTkFrame(self.content, corner_radius=18, fg_color=self.theme["surface_alt"])
        self.progress_frame.grid(row=1, column=0, padx=24, pady=(0, 12), sticky="ew")
        self.progress_frame.grid_columnconfigure(0, weight=1)

        self.status_label = ctk.CTkLabel(
            self.progress_frame,
            text="Selecione um banco SQLite ou Firebird para iniciar no CONVERSOR - VEXPER.",
            anchor="w",
            text_color="#183A5A",
        )
        self.status_label.grid(row=0, column=0, padx=18, pady=(14, 8), sticky="ew")

        self.progress_bar = ctk.CTkProgressBar(self.progress_frame, progress_color="#2A9D8F")
        self.progress_bar.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="ew")
        self.progress_bar.set(0)

        self.action_frame = ctk.CTkFrame(self.content, corner_radius=18, fg_color=self.theme["surface"])
        self.action_frame.grid(row=2, column=0, padx=24, pady=(0, 12), sticky="ew")
        self.action_frame.grid_columnconfigure(0, weight=1)
        self.action_frame.grid_rowconfigure(0, weight=1)

        self.selected_table_label = ctk.CTkLabel(
            self.action_frame,
            text="Tabela selecionada: nenhuma",
            anchor="w",
            text_color="#183A5A",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
        )
        self.selected_table_label.grid(row=0, column=0, padx=18, pady=16, sticky="ew")

        self.export_button = ctk.CTkButton(
            self.action_frame,
            text="Converter tabela selecionada para Excel",
            height=42,
            corner_radius=14,
            fg_color="#183A5A",
            hover_color="#102A43",
            text_color="#FFFFFF",
            command=self.export_selected_tables,
            state="disabled",
            width=280,
        )
        self.export_button.grid(row=0, column=1, padx=18, pady=12, sticky="e")

        self.preview_card = ctk.CTkFrame(self.content, corner_radius=20, fg_color=self.theme["surface"])
        self.preview_card.grid(row=3, column=0, padx=24, pady=(0, 24), sticky="nsew")
        self.preview_card.grid_columnconfigure(0, weight=1)
        self.preview_card.grid_rowconfigure(1, weight=1)

        self.preview_title = ctk.CTkLabel(
            self.preview_card,
            text="Pré-visualização",
            font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"),
            text_color="#183A5A",
        )
        self.preview_title.grid(row=0, column=0, padx=20, pady=(18, 8), sticky="w")

        self.preview_card.grid_columnconfigure(1, weight=0)

        self.preview_frame = tk.Frame(self.preview_card, bg=self.theme["surface"])
        self.preview_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=(0, 20), sticky="nsew")
        self.preview_frame.grid_columnconfigure(0, weight=1)
        self.preview_frame.grid_rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(self.preview_frame, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")

        self.preview_vertical_scrollbar = tk.Scrollbar(
            self.preview_frame,
            orient="vertical",
            command=self.tree.yview,
            width=18,
            activebackground=self.theme["accent"],
            bg=self.theme["surface_alt"],
            troughcolor=self.theme["app_bg"],
            relief="flat",
            bd=0,
        )
        self.preview_vertical_scrollbar.grid(row=0, column=1, sticky="ns")

        self.preview_horizontal_scrollbar = tk.Scrollbar(
            self.preview_frame,
            orient="horizontal",
            command=self.tree.xview,
            width=18,
            activebackground=self.theme["accent"],
            bg=self.theme["surface_alt"],
            troughcolor=self.theme["app_bg"],
            relief="flat",
            bd=0,
        )
        self.preview_horizontal_scrollbar.grid(row=1, column=0, sticky="ew")

        self.tree.configure(
            yscrollcommand=self.preview_vertical_scrollbar.set,
            xscrollcommand=self.preview_horizontal_scrollbar.set,
        )
        self.tree.bind("<MouseWheel>", self._on_tree_mousewheel)
        self.tree.bind("<Shift-MouseWheel>", self._on_tree_horizontal_mousewheel)

        self.tree_style = ttk.Style()
        self.tree_style.theme_use("default")
        self.tree_style.configure("Treeview", rowheight=28, font=("Segoe UI", 10), background=self.theme["list_bg"])
        self.tree_style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

    def _on_listbox_mousewheel(self, event) -> str:
        self.table_listbox.yview_scroll(int(-event.delta / 120), "units")
        return "break"

    def _on_tree_mousewheel(self, event) -> str:
        self.tree.yview_scroll(int(-event.delta / 120), "units")
        return "break"

    def _on_tree_horizontal_mousewheel(self, event) -> str:
        self.tree.xview_scroll(int(-event.delta / 120), "units")
        return "break"

    def _toggle_connection_panel(self) -> None:
        self.connection_panel_visible = not self.connection_panel_visible
        if self.connection_panel_visible:
            self.connection_frame.pack(fill="x", padx=22, pady=(0, 14))
            self.connection_toggle_button.configure(text="Ocultar conexão Firebird")
        else:
            self.connection_frame.pack_forget()
            self.connection_toggle_button.configure(text="Mostrar conexão Firebird")

    def _sidebar_entry(self, label: str, value: str, parent=None, show: str | None = None):
        container = parent or self.sidebar
        ctk.CTkLabel(
            container,
            text=label,
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color="#486581",
        ).pack(anchor="w", padx=16 if parent else 22, pady=(0, 4))

        entry = ctk.CTkEntry(container, height=32)
        entry.pack(fill="x", padx=16 if parent else 22, pady=(0, 8))
        entry.insert(0, value)
        if show is not None:
            entry.configure(show=show)
        return entry

    def _build_converter(self) -> DatabaseExcelConverter:
        return DatabaseExcelConverter(
            self.database_path,
            host=self.host_entry.get().strip() or "127.0.0.1",
            port=int(self.port_entry.get().strip() or "3050"),
            user=self.user_entry.get().strip() or "SYSDBA",
            password=self.password_entry.get(),
            charset=self.charset_entry.get().strip() or "WIN1252",
        )

    def _metric_card(self, parent, label: str, value: str, column: int):
        card = ctk.CTkFrame(parent, corner_radius=18, fg_color=self.theme["surface"])
        card.grid(row=0, column=column, padx=8, sticky="ew")

        ctk.CTkLabel(
            card,
            text=label,
            font=ctk.CTkFont(family="Segoe UI", size=14),
            text_color="#486581",
        ).pack(anchor="w", padx=18, pady=(16, 2))

        value_label = ctk.CTkLabel(
            card,
            text=value,
            font=ctk.CTkFont(family="Segoe UI", size=28, weight="bold"),
            text_color="#102A43",
        )
        value_label.pack(anchor="w", padx=18, pady=(0, 16))
        return value_label

    def pick_database(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Selecione um banco de dados",
            filetypes=[("Bancos suportados", "*.db *.sqlite *.sqlite3 *.fdb"), ("Todos os arquivos", "*.*")],
        )
        if not file_path:
            return

        self.database_path = Path(file_path)
        self.converter = None
        self.file_label.configure(text=str(self.database_path))
        self.scan_button.configure(state="normal")
        self.export_all_button.configure(state="disabled")
        self.export_button.configure(state="disabled")
        self.select_all_button.configure(state="disabled")
        self.all_tables.clear()
        self.table_search_var.set("")
        self.table_listbox.delete(0, "end")
        self.filtered_tables.clear()
        self.table_profiles.clear()
        self._update_metrics(0, 0, 0)
        self.status_label.configure(text="Arquivo selecionado. Lendo estrutura do banco para listar as tabelas...")
        self.progress_bar.set(0.05)
        self._clear_preview()
        if bool(self.preferences.get("auto_scan", True)):
            self.scan_database()

    def scan_database(self) -> None:
        if not self.database_path:
            return

        self.scan_button.configure(state="disabled")
        self.export_all_button.configure(state="disabled")
        self.export_button.configure(state="disabled")
        self.pick_button.configure(state="disabled")
        self.select_all_button.configure(state="disabled")
        self.status_label.configure(text="Lendo estrutura do banco em segundo plano...")
        self.progress_bar.set(0.1)
        self.table_listbox.delete(0, "end")
        self.all_tables.clear()
        self.filtered_tables.clear()
        self.table_profiles.clear()
        self._update_metrics(0, 0, 0)
        self._clear_preview()

        try:
            converter = self._build_converter()
        except (TypeError, ValueError) as error:
            self._restore_actions()
            messagebox.showerror("Configuração inválida", str(error))
            return

        self.scan_thread = threading.Thread(target=self._run_scan, args=(converter,), daemon=True)
        self.scan_thread.start()

    def _run_scan(self, converter: DatabaseExcelConverter) -> None:
        try:
            tables = converter.list_tables()
            self.progress_queue.put(("scan_done", (converter, tables)))
        except Exception as error:
            self.progress_queue.put(("scan_error", str(error)))

    def select_all_tables(self) -> None:
        if self.table_listbox.size() > 0:
            self.table_listbox.selection_clear(0, "end")
            self.table_listbox.selection_set(0)
            self.on_table_selected(None)

    def _on_table_search_changed(self, *_args) -> None:
        self._apply_table_filter()

    def _apply_table_filter(self, preserve_view: bool = True, auto_select_first: bool = True) -> None:
        query = self.table_search_var.get().strip().upper()
        selected_names = set(self._selected_tables())
        current_yview = self.table_listbox.yview()[0] if self.table_listbox.size() > 0 else 0.0

        self.table_listbox.delete(0, "end")
        self.filtered_tables = []
        for table_name in self.all_tables:
            if query and query not in table_name.upper():
                continue
            self.filtered_tables.append(table_name)
            self.table_listbox.insert("end", self._format_table_label(table_name))

        for index in range(self.table_listbox.size()):
            if self.filtered_tables[index] in selected_names:
                self.table_listbox.selection_set(index)
                break

        if self.table_listbox.size() > 0 and auto_select_first and not self.table_listbox.curselection():
            self.table_listbox.selection_set(0)

        if preserve_view and self.table_listbox.size() > 0:
            self.table_listbox.yview_moveto(current_yview)

        if self.table_listbox.size() == 0:
            self.status_label.configure(text="Nenhuma tabela encontrada com esse filtro.")

    def _format_table_label(self, table_name: str) -> str:
        profile = self.table_profiles.get(table_name)
        if profile is None:
            return f"{table_name} | carregando..."
        return f"{table_name} | {profile.row_count:,} linha(s)".replace(",", ".")

    def on_table_selected(self, _event) -> None:
        selected = self._selected_tables()
        if not selected or not self.converter:
            self.selected_table_label.configure(text="Tabela selecionada: nenhuma")
            return

        table_name = selected[0]
        self.selected_table_label.configure(text=f"Tabela selecionada: {table_name}")
        self.status_label.configure(text=f"Carregando pré-visualização de {table_name}...")
        self.progress_bar.set(0.24)
        self.preview_thread = threading.Thread(target=self._run_preview, args=(table_name,), daemon=True)
        self.preview_thread.start()

    def _run_preview(self, table_name: str) -> None:
        try:
            if not self.converter:
                return
            columns, rows = self.converter.read_preview(table_name)
            profile = self.table_profiles.get(table_name)
            if profile is None or not profile.columns:
                profile = self.converter.get_profile(table_name)
            self.progress_queue.put(("preview_done", (table_name, columns, rows, profile)))
        except Exception as error:
            self.progress_queue.put(("preview_error", str(error)))

    def export_selected_tables(self) -> None:
        if not self.converter or not self.database_path:
            return

        selected_tables = self._selected_tables()
        if not selected_tables:
            messagebox.showwarning("Seleção necessária", "Selecione uma tabela para converter.")
            return

        table_name = selected_tables[0]
        has_template = template_path_for_table(table_name) is not None
        suggested_name = f"{table_name}.xls" if has_template else f"{table_name}.xlsx"
        output_path = self._ask_output_path(suggested_name, prefer_xls=has_template)
        if output_path is None:
            return

        self._start_export([table_name], output_path)

    def export_all_tables(self) -> None:
        if not self.converter or not self.database_path:
            return

        table_names = list(self.all_tables)
        if not table_names:
            messagebox.showwarning("Sem tabelas", "Nenhuma tabela disponível para exportação.")
            return

        output_path = self._ask_output_path(f"{self.database_path.stem}_completo.xlsx", prefer_xls=False)
        if output_path is None:
            return

        self._start_export(table_names, output_path)

    def _ask_output_path(self, suggested_name: str, prefer_xls: bool = False) -> Path | None:
        output_path = filedialog.asksaveasfilename(
            title="Salvar arquivo Excel",
            defaultextension=".xls" if prefer_xls else ".xlsx",
            initialfile=suggested_name,
            filetypes=[("Excel 97-2003", "*.xls"), ("Excel moderno", "*.xlsx")],
        )
        if not output_path:
            return None
        return Path(output_path)

    def _start_export(self, selected_tables: list[str], output_path: Path) -> None:
        export_mode = "banco inteiro" if len(selected_tables) == len(self.all_tables) else selected_tables[0]
        self.export_all_button.configure(state="disabled")
        self.export_button.configure(state="disabled")
        self.scan_button.configure(state="disabled")
        self.pick_button.configure(state="disabled")
        self.select_all_button.configure(state="disabled")
        self.progress_bar.set(0.32)
        self.status_label.configure(text=f"Iniciando exportação de {export_mode} em {output_path.name}...")

        self.export_thread = threading.Thread(
            target=self._run_export,
            args=(selected_tables, output_path),
            daemon=True,
        )
        self.export_thread.start()

    def _run_export(self, selected_tables: list[str], output_path: Path) -> None:
        try:
            if not self.converter:
                return
            profiles = self.converter.export_tables(selected_tables, output_path, self._push_progress)
            self.progress_queue.put(("done", ExportResult(output_path=output_path, profiles=profiles)))
        except Exception as error:
            self.progress_queue.put(("error", str(error)))

    def _push_progress(self, message: str, ratio: float) -> None:
        self.progress_queue.put(("progress", (message, ratio)))

    def _poll_queue(self) -> None:
        while True:
            try:
                kind, payload = self.progress_queue.get_nowait()
            except queue.Empty:
                break

            if kind == "progress":
                message, ratio = payload
                self.status_label.configure(text=message)
                self.progress_bar.set(ratio)
            elif kind == "update_progress":
                self.status_label.configure(text=str(payload))
            elif kind == "update_ready":
                package = payload
                self.pending_update = package
                self.status_label.configure(text=f"Atualizacao {package.version} identificada. Escolha quando deseja instalar.")
                self._prompt_update_package(package)
            elif kind == "update_error":
                # Falha silenciosa para nao atrapalhar o uso do sistema.
                pass
            elif kind == "scan_done":
                self.converter, tables = payload
                self.all_tables = list(tables)
                self._apply_table_filter(preserve_view=False, auto_select_first=True)

                self.profile_thread = threading.Thread(target=self._load_table_profiles, args=(list(tables),), daemon=True)
                self.profile_thread.start()

                self._update_metrics(len(tables), 0, 0)
                self.pick_button.configure(state="normal")
                self.scan_button.configure(state="normal")
                self.select_all_button.configure(state="normal" if tables else "disabled")
                self.export_all_button.configure(state="normal" if tables else "disabled")
                self.export_button.configure(state="normal" if tables else "disabled")
                self.progress_bar.set(0.2 if tables else 0)

                if tables:
                    self.status_label.configure(
                        text=f"{len(tables)} tabela(s) encontradas. Selecione uma tabela para visualizar e converter para Excel."
                    )
                    self.table_listbox.selection_set(0)
                    self.selected_table_label.configure(text=f"Tabela selecionada: {tables[0]}")
                    self.on_table_selected(None)
                else:
                    self._clear_preview()
                    self.selected_table_label.configure(text="Tabela selecionada: nenhuma")
                    self.status_label.configure(text="Nenhuma tabela encontrada no banco.")
                    messagebox.showwarning("Sem tabelas", "O banco foi aberto, mas não contém tabelas para exportação.")
            elif kind == "scan_error":
                self._restore_actions()
                self.status_label.configure(text="Falha ao ler a estrutura do banco.")
                self.progress_bar.set(0)
                messagebox.showerror("Falha na leitura", str(payload))
            elif kind == "preview_done":
                table_name, columns, rows, profile = payload
                self.table_profiles[table_name] = profile
                self._apply_table_filter(preserve_view=True, auto_select_first=False)
                self._populate_preview(table_name, columns, rows)
                self.metric_rows.configure(text=f"{profile.row_count:,}".replace(",", "."))
                self.metric_columns.configure(text=f"{profile.column_count:,}".replace(",", "."))
                self.status_label.configure(
                    text=f"Tabela {table_name}: {profile.row_count} linha(s), {profile.column_count} coluna(s)."
                )
                self.progress_bar.set(0.32)
            elif kind == "profile_loaded":
                profile = payload
                self.table_profiles[profile.name] = profile
                self._apply_table_filter(preserve_view=True, auto_select_first=False)
            elif kind == "preview_error":
                self.status_label.configure(text="Falha ao carregar a pré-visualização.")
                self.progress_bar.set(0)
                messagebox.showerror("Erro na pré-visualização", str(payload))
            elif kind == "done":
                total_rows = sum(profile.row_count for profile in payload.profiles)
                total_columns = sum(profile.column_count for profile in payload.profiles)
                self._update_metrics(len(payload.profiles), total_rows, total_columns)
                self.status_label.configure(text=f"Excel gerado com sucesso em {payload.output_path}")
                self.progress_bar.set(1)
                self._restore_actions()
                if bool(self.preferences.get("sound_enabled", True)):
                    play_completion_sound()
                messagebox.showinfo("Conversão concluída", f"Arquivo salvo em:\n{payload.output_path}")
            elif kind == "error":
                self.status_label.configure(text="Falha ao gerar o Excel.")
                self.progress_bar.set(0)
                self._restore_actions()
                messagebox.showerror("Erro na exportação", str(payload))

        self.after(120, self._poll_queue)

    def _restore_actions(self) -> None:
        self.pick_button.configure(state="normal")
        self.scan_button.configure(state="normal" if self.database_path else "disabled")
        has_tables = len(self.all_tables) > 0
        self.select_all_button.configure(state="normal" if has_tables else "disabled")
        self.export_all_button.configure(state="normal" if has_tables else "disabled")
        self.export_button.configure(state="normal" if has_tables else "disabled")

    def _selected_tables(self) -> list[str]:
        indices = self.table_listbox.curselection()
        return [self.filtered_tables[index] for index in indices if index < len(self.filtered_tables)]

    def _load_table_profiles(self, table_names: list[str]) -> None:
        if not self.converter:
            return
        for table_name in table_names:
            if table_name in self.table_profiles:
                continue
            try:
                profile = self.converter.get_profile(table_name)
                self.progress_queue.put(("profile_loaded", profile))
            except Exception:
                continue

    def _update_metrics(self, table_count: int, row_count: int, column_count: int) -> None:
        self.metric_tables.configure(text=f"{table_count:,}".replace(",", "."))
        self.metric_rows.configure(text=f"{row_count:,}".replace(",", "."))
        self.metric_columns.configure(text=f"{column_count:,}".replace(",", "."))

    def _clear_preview(self) -> None:
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()
        self.preview_title.configure(text="Pré-visualização")
        self.selected_table_label.configure(text="Tabela selecionada: nenhuma")

    def _populate_preview(self, table_name: str, columns: list[str], rows: list[tuple]) -> None:
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = columns

        for column in columns:
            self.tree.heading(column, text=column)
            width = min(max(len(column) * 11, 120), 260)
            self.tree.column(column, width=width, minwidth=120, stretch=False, anchor="w")

        for row in rows:
            self.tree.insert("", "end", values=[normalize_cell_value(value) for value in row])

        self.preview_title.configure(text=f"Pré-visualização: {table_name}")

    def _apply_theme(self) -> None:
        self.configure(fg_color=self.theme["app_bg"])
        self.header.configure(fg_color=self.theme["header"])
        self.badge_label.configure(text_color=self.theme["badge"], fg_color=self.theme["header_alt"])
        self.version_label.configure(text_color="#DCE6F2", fg_color=self.theme["header_alt"])
        self.sidebar.configure(fg_color=self.theme["surface"])
        self.content.configure(fg_color=self.theme["surface"])
        self.list_frame.configure(bg=self.theme["surface"])
        self.table_listbox.configure(bg=self.theme["list_bg"], fg=self.theme["text"], selectbackground=self.theme["accent"], selectforeground="#FFFFFF")
        self.table_scrollbar.configure(activebackground=self.theme["accent"], bg=self.theme["surface_alt"], troughcolor=self.theme["app_bg"])
        self.progress_frame.configure(fg_color=self.theme["surface_alt"])
        self.status_label.configure(text_color=self.theme["text"])
        self.progress_bar.configure(progress_color=self.theme["accent"])
        self.action_frame.configure(fg_color=self.theme["surface"])
        self.selected_table_label.configure(text_color=self.theme["text"])
        self.export_button.configure(fg_color=self.theme["secondary"], hover_color=self.theme["secondary_hover"])
        self.pick_button.configure(fg_color=self.theme["accent"], hover_color=self.theme["accent_hover"])
        self.connection_toggle_button.configure(fg_color=self.theme["surface_alt"], hover_color=self.theme["app_bg"], text_color=self.theme["text"])
        self.connection_frame.configure(fg_color=self.theme["surface_alt"])
        self.scan_button.configure(fg_color=self.theme["secondary"], hover_color=self.theme["secondary_hover"])
        self.select_all_button.configure(fg_color=self.theme["surface_alt"], hover_color=self.theme["app_bg"], text_color=self.theme["text"])
        self.export_all_button.configure(fg_color=self.theme["warm"], hover_color=self.theme["warm_hover"], text_color="#1F2933")
        self.preview_card.configure(fg_color=self.theme["surface"])
        self.preview_frame.configure(bg=self.theme["surface"])
        self.preview_title.configure(text_color=self.theme["text"])
        self.preview_vertical_scrollbar.configure(activebackground=self.theme["accent"], bg=self.theme["surface_alt"], troughcolor=self.theme["app_bg"])
        self.preview_horizontal_scrollbar.configure(activebackground=self.theme["accent"], bg=self.theme["surface_alt"], troughcolor=self.theme["app_bg"])
        self.settings_button.configure(fg_color=self.theme["surface_alt"], text_color=self.theme["text"], hover_color=self.theme["app_bg"])
        for metric in (self.metric_tables, self.metric_rows, self.metric_columns):
            metric.configure(text_color=self.theme["secondary"])
        self.tree_style.configure(
            "Treeview",
            rowheight=28,
            font=("Segoe UI", 10),
            background=self.theme["list_bg"],
            fieldbackground=self.theme["list_bg"],
            foreground=self.theme["text"],
        )
        self.tree_style.map("Treeview", background=[("selected", self.theme["accent"])], foreground=[("selected", "#FFFFFF")])
        self.tree_style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background=self.theme["surface_alt"], foreground=self.theme["text"])

    def open_settings(self) -> None:
        if self.settings_window is not None and self.settings_window.winfo_exists():
            self.settings_window.focus()
            return

        self.settings_window = ctk.CTkToplevel(self)
        self.settings_window.title("Configuracao")
        self.settings_window.geometry("560x560+520+170")
        self.settings_window.resizable(False, False)
        self.settings_window.configure(fg_color=self.theme["surface"])

        ctk.CTkLabel(
            self.settings_window,
            text="Configuracoes do sistema",
            font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"),
            text_color=self.theme["text"],
        ).pack(anchor="w", padx=24, pady=(22, 8))

        ctk.CTkLabel(self.settings_window, text="Tema visual", text_color=self.theme["muted"]).pack(anchor="w", padx=24)
        theme_var = tk.StringVar(value=self.theme_name)
        theme_menu = ctk.CTkOptionMenu(self.settings_window, values=list(THEMES.keys()), variable=theme_var)
        theme_menu.pack(fill="x", padx=24, pady=(6, 16))

        sound_var = tk.BooleanVar(value=bool(self.preferences.get("sound_enabled", True)))
        auto_scan_var = tk.BooleanVar(value=bool(self.preferences.get("auto_scan", True)))
        remember_var = tk.BooleanVar(value=bool(self.preferences.get("remember_login", False)))
        auto_update_var = tk.BooleanVar(value=bool(self.preferences.get("auto_update_enabled", True)))
        update_feed_var = tk.StringVar(value=str(self.preferences.get("update_feed", default_update_feed())))

        for text, variable in (
            ("Tocar som ao finalizar exportacao", sound_var),
            ("Ler estrutura automaticamente ao escolher o banco", auto_scan_var),
            ("Manter login salvo neste computador", remember_var),
            ("Verificar atualizacoes ao abrir o sistema", auto_update_var),
        ):
            ctk.CTkCheckBox(
                self.settings_window,
                text=text,
                variable=variable,
                text_color=self.theme["text"],
                fg_color=self.theme["accent"],
                hover_color=self.theme["accent_hover"],
            ).pack(anchor="w", padx=24, pady=8)

        ctk.CTkLabel(
            self.settings_window,
            text="Pasta ou URL do feed de atualizacao",
            text_color=self.theme["muted"],
        ).pack(anchor="w", padx=24, pady=(14, 6))

        ctk.CTkEntry(
            self.settings_window,
            textvariable=update_feed_var,
            height=38,
        ).pack(fill="x", padx=24)

        ctk.CTkLabel(
            self.settings_window,
            text="Temas disponiveis: Oceano, Grafite, Aurora e Solar. Cada um muda a paleta e o contraste do layout. Para atualizar todos os computadores, use uma pasta compartilhada ou uma URL do GitHub Releases, por exemplo https://github.com/usuario/repositorio/releases/latest/download/ .",
            text_color=self.theme["muted"],
            wraplength=420,
            justify="left",
        ).pack(anchor="w", padx=24, pady=(18, 20))

        def save_and_apply() -> None:
            self.preferences["theme"] = theme_var.get()
            self.preferences["sound_enabled"] = sound_var.get()
            self.preferences["auto_scan"] = auto_scan_var.get()
            self.preferences["remember_login"] = remember_var.get()
            self.preferences["auto_update_enabled"] = auto_update_var.get()
            self.preferences["update_feed"] = update_feed_var.get().strip()
            if not remember_var.get():
                self.preferences["saved_username"] = ""
                self.preferences["saved_password"] = ""
            save_preferences(self.preferences)
            self.theme_name = str(self.preferences["theme"])
            self.theme = THEMES.get(self.theme_name, THEMES["Oceano"])
            self._apply_theme()
            self.settings_window.destroy()

        ctk.CTkButton(
            self.settings_window,
            text="Salvar preferencias",
            height=42,
            fg_color=self.theme["accent"],
            hover_color=self.theme["accent_hover"],
            command=save_and_apply,
        ).pack(fill="x", padx=24, pady=(0, 10))

        ctk.CTkButton(
            self.settings_window,
            text="Fechar",
            height=40,
            fg_color=self.theme["surface_alt"],
            text_color=self.theme["text"],
            hover_color=self.theme["app_bg"],
            command=self.settings_window.destroy,
        ).pack(fill="x", padx=24)


if __name__ == "__main__":
    app = App()
    app.mainloop()